"""
Motor de imputación. Importable desde app.py (Streamlit) o scripts CLI.
No tiene side effects al importar.
"""

import io
import re
import datetime
import openpyxl
from openpyxl.styles import PatternFill

from mep_helper import mep_para_fecha

YELLOW_FILL = PatternFill(patternType='solid', fgColor='FFFF00')

MESES_ES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
    5: 'mayo',  6: 'junio',   7: 'julio', 8: 'agosto',
    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre',
}

SHEETS_BASE_PESOS = {
    'INDICE CAC': {
        'header_row': 5, 'data_start': 6,
        'cuit_col': 12, 'nombre_col': 9, 'lote_col': 8,
    },
    'INDICE CAC M. OBRA': {
        'header_row': 3, 'data_start': 4,
        'cuit_col': 11, 'nombre_col': 9, 'lote_col': 8,
    },
    'BOLSA CEMENTO': {
        'header_row': 3, 'data_start': 4,
        'cuit_col': 14, 'nombre_col': 10, 'lote_col': 9,
    },
}

SHEETS_BASE_USD = {
    '$  USD fijo': {
        'header_row': 3, 'data_start': 4,
        'cuit_col': 13, 'nombre_col': 10, 'lote_col': 9,
    },
}


def normalize_cuits(raw):
    """Extrae todos los CUITs de una celda de deudores.

    Maneja guiones/puntos (20-34658691-6), varios CUITs separados por '/'
    y celdas con texto libre entre CUITs
    ('27-14742775-7 (esther) Eluen 20-44090800-5' → dos CUITs).
    """
    if raw is None or raw == '':
        return []
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = str(raw)
    # Unir grupos de dígitos conectados por guion/punto (con espacios sueltos
    # alrededor) mientras no excedan los 12 dígitos de un CUIT. Un guion entre
    # dos CUITs completos ("20259477895 - 27267941594") no los une porque el
    # resultado se pasaría de largo.
    results, cur, prev_end = [], '', None
    for g in re.finditer(r'\d+', s):
        sep = s[prev_end:g.start()] if prev_end is not None else None
        unido = sep is not None and re.fullmatch(r'\s*[.\-]\s*', sep)
        if cur and unido and len(cur) + len(g.group()) <= 12:
            cur += g.group()
        else:
            if 10 <= len(cur) <= 12:
                results.append(cur)
            cur = g.group()
        prev_end = g.end()
    if 10 <= len(cur) <= 12:
        results.append(cur)
    if not results:
        # formato raro (ej: dígitos separados solo por espacios): unir todo
        for part in s.split('/'):
            digits = re.sub(r'\D', '', part)
            if len(digits) >= 10:
                results.append(digits)
    return list(dict.fromkeys(results))


def is_row_yellow(ws, row_num):
    for cell in ws[row_num]:
        if cell.value is None:
            continue
        fill = cell.fill
        if fill and fill.patternType and fill.patternType != 'none':
            rgb = str(fill.fgColor.rgb) if fill.fgColor else ''
            if 'FFFF00' in rgb or rgb == 'FFFFFF00':
                return True
        break
    return False


def extract_cuit_from_concepto(concepto):
    if not concepto:
        return None
    matches = re.findall(r'\b(\d{11,12})\b', str(concepto))
    return matches[0] if matches else None


def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, str):
        for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


def _norm(s):
    return s.lower().replace('ó', 'o').replace('é', 'e').replace('á', 'a').replace('í', 'i').replace('ú', 'u')


def max_cuota_celda(val):
    """Máxima cuota implicada por una celda de "NUMERO DE CUOTA".

    Acepta números, listas tipo "10 y 11" / "2 Y 3" / "3, 4 y 5" (el formato que
    escribe el propio bot al imputar varias cuotas juntas) y "parte de cX"
    (la cuota X está parcialmente paga → el siguiente pago completo es X+1).
    Otros textos ("1 al 12 inclusive", etc.) se ignoran por ambiguos.
    """
    if isinstance(val, (int, float)):
        n = int(val)
        return n if 0 < n <= 200 else None
    if isinstance(val, str):
        s = val.strip()
        if 'parte' in s.lower():
            m = re.search(r'\d+', s)
            nums = [int(m.group())] if m else []
        elif re.fullmatch(r'[cC]?\s*\d+(\s*[,yY]\s*[cC]?\s*\d+)*\.?', s):
            nums = [int(x) for x in re.findall(r'\d+', s)]
        else:
            return None
        nums = [n for n in nums if 0 < n <= 200]
        return max(nums) if nums else None
    return None


def detectar_mes_transferencias(ws_imp, max_row=500):
    """Lee col A de imputaciones y retorna (year, month) más frecuente, o (None, None).

    Solo cuenta filas PENDIENTES (no amarillas): las hojas —sobre todo las USD—
    son acumulativas y arrastran transferencias ya imputadas de meses anteriores.
    Contar todas hacía que el mes detectado fuera un mes viejo ya imputado y el
    script reportara "mes ya imputado" para todo sin imputar nada.
    Si no hubiera filas pendientes, cae a contar todas (comportamiento previo).
    """
    from collections import Counter
    conteo = Counter()
    conteo_todas = Counter()
    for row in ws_imp.iter_rows(min_row=4, max_row=max_row, max_col=1):
        dt = parse_date(row[0].value)
        if not dt:
            continue
        conteo_todas[(dt.year, dt.month)] += 1
        if not is_row_yellow(ws_imp, row[0].row):
            conteo[(dt.year, dt.month)] += 1
    if not conteo:
        conteo = conteo_todas
    if not conteo:
        return None, None
    (year, month), _ = conteo.most_common(1)[0]
    return year, month


def detectar_columnas_mes(ws, header_row, year=None, month=None):
    """
    Retorna la columna 'teorico' para el mes/año indicado.
    Si year/month son None o no hay match, retorna la más a la derecha (fallback).
    """
    mes_nombre = _norm(MESES_ES.get(month, '')) if month else None
    yr_str = str(year % 100) if year else None       # "26" para 2026
    yr_full = str(year) if year else None             # "2026"

    teo_col_match = None
    teo_col_fallback = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if not h:
            continue
        h_norm = _norm(str(h))
        if 'teorico' not in h_norm:
            continue
        teo_col_fallback = c
        if mes_nombre and yr_str and mes_nombre in h_norm:
            if yr_str in h_norm or (yr_full and yr_full in h_norm):
                teo_col_match = c
    return teo_col_match if teo_col_match is not None else teo_col_fallback


def _col_por_header(ws, header_row, keywords):
    """Primera columna cuyo header (normalizado) contiene alguna de keywords."""
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if not h:
            continue
        hn = _norm(str(h))
        if any(k in hn for k in keywords):
            return c
    return None


def build_sheets_cfg(wb_deu_data, sheets_base, year=None, month=None):
    """Auto-detecta columnas del mes y construye sheets_cfg completo."""
    sheets_cfg = {}
    mes_info = {}
    for sheet_name, base in sheets_base.items():
        try:
            ws = wb_deu_data[sheet_name]
        except KeyError:
            continue
        teo_col = detectar_columnas_mes(ws, base['header_row'], year=year, month=month)
        if teo_col is None:
            continue
        cfg = dict(base)
        # Auto-detectar CUIT/Nombre/LOTE por header (los índices fijos se rompen
        # si el archivo de deudores agrega/quita columnas). Se cae al valor base
        # si no se encuentra el header.
        hr = base['header_row']
        cuit_c   = _col_por_header(ws, hr, ('cuit',))
        nombre_c = _col_por_header(ws, hr, ('nombre',))
        lote_c   = _col_por_header(ws, hr, ('lote',))
        if cuit_c:
            cfg['cuit_col'] = cuit_c
        if nombre_c:
            cfg['nombre_col'] = nombre_c
        if lote_c:
            cfg['lote_col'] = lote_c
        cfg['teo_col'] = teo_col
        cfg['real_col'] = teo_col + 1
        cfg['cuota_col'] = teo_col + 2
        cfg['fecha_col'] = teo_col + 3
        cfg['max_cuota_col'] = 1
        sheets_cfg[sheet_name] = cfg
        header_val = ws.cell(base['header_row'], teo_col).value
        mes_info[sheet_name] = str(header_val).strip() if header_val else f'col {teo_col}'
    return sheets_cfg, mes_info


def build_indices(wb_deu_data, sheets_cfg):
    cuit_index = {}
    nombre_index = {}
    cuota_history_cols = {}

    for sheet_name, cfg in sheets_cfg.items():
        ws_data = wb_deu_data[sheet_name]
        max_row = ws_data.max_row

        for r in range(cfg['data_start'], max_row + 1):
            nombre = ws_data.cell(r, cfg['nombre_col']).value
            cuit_raw = ws_data.cell(r, cfg['cuit_col']).value
            if not nombre and not cuit_raw:
                continue
            for c in normalize_cuits(cuit_raw):
                cuit_index.setdefault(c, []).append((sheet_name, r, nombre))

        for r in range(cfg['data_start'], max_row + 1):
            nombre = ws_data.cell(r, cfg['nombre_col']).value
            if not nombre:
                continue
            for palabra in str(nombre).upper().split():
                if len(palabra) >= 4:
                    # normalizar tildes para que "ÁNGEL" y "ANGEL" sean la misma clave
                    nombre_index.setdefault(_norm(palabra).upper(), []).append((sheet_name, r, nombre))

        cols = []
        for c in range(1, ws_data.max_column + 1):
            h = ws_data.cell(cfg['header_row'], c).value
            if h and ('numero' in str(h).lower() or 'n°' in str(h).lower() or 'nro' in str(h).lower()) and 'cuota' in str(h).lower():
                if c != cfg['cuota_col']:
                    cols.append(c)
        cuota_history_cols[sheet_name] = cols

    return cuit_index, nombre_index, cuota_history_cols


def buscar_en_deudores_por_nombre(nombre_str, nombre_index):
    from collections import Counter

    # Si tiene "/" (ej: "Cerda Gabriel / Iboldi Yanina"), buscar cada parte por separado
    if '/' in nombre_str:
        resultado, seen = [], set()
        for parte in nombre_str.split('/'):
            for item in buscar_en_deudores_por_nombre(parte.strip(), nombre_index):
                k = (item[0], item[1])
                if k not in seen:
                    resultado.append(item)
                    seen.add(k)
        return resultado

    # Normalizar tildes en las palabras buscadas (igual que en el índice)
    palabras = [_norm(p).upper() for p in nombre_str.split() if len(p) >= 4]
    if not palabras:
        return []
    sets = [set((s, r) for s, r, _ in nombre_index.get(p, [])) for p in palabras]
    if not sets:
        return []
    comunes = sets[0]
    for s in sets[1:]:
        comunes &= s
    if not comunes:
        cnt = Counter()
        for p in palabras:
            for s, r, _ in nombre_index.get(p, []):
                cnt[(s, r)] += 1
        max_hits = max(cnt.values()) if cnt else 0
        comunes = {k for k, v in cnt.items() if v == max_hits and v >= max(2, len(palabras) - 1)}
    resultado = []
    seen = set()
    for p in palabras:
        for s, r, n in nombre_index.get(p, []):
            if (s, r) in comunes and (s, r) not in seen:
                resultado.append((s, r, n))
                seen.add((s, r))
    return resultado


def build_previo(wb_imp, imp_sheet, es_usd=False):
    all_sheets = wb_imp.sheetnames
    try:
        idx_actual = all_sheets.index(imp_sheet)
        previas = list(reversed(all_sheets[:idx_actual]))
        if es_usd:
            previas = [h for h in previas if 'usd' in h.lower()][:6]
        else:
            previas = previas[:8]
    except ValueError:
        previas = []

    cuit_to_nombre_previo = {}
    cuit_to_cuota_previo = {}
    for hoja in previas:
        try:
            ws_prev = wb_imp[hoja]
        except Exception:
            continue
        for row in ws_prev.iter_rows(min_row=4, max_row=ws_prev.max_row):
            concepto = row[2].value if len(row) > 2 else None
            col_h = row[7].value if len(row) > 7 else None
            if not concepto or not col_h:
                continue
            col_h_str = str(col_h).strip()
            if not col_h_str or 'PAGO MENOS' in col_h_str or 'Saldo' in col_h_str:
                continue
            cuit = extract_cuit_from_concepto(concepto)
            if cuit and cuit not in cuit_to_nombre_previo:
                # Maneja "c21", "c21,22 y 23", etc.
                m_cuota = re.search(r'c(\d+(?:[\s,y]+\d+)*)\s*$', col_h_str, re.IGNORECASE)
                if m_cuota:
                    nums = re.findall(r'\d+', m_cuota.group(1))
                    if nums:
                        cuit_to_cuota_previo[cuit] = max(int(n) for n in nums)
                nombre_prev = re.sub(r'\s*c\d+(?:[\s,y]+\d+)*\s*$', '', col_h_str, flags=re.IGNORECASE).strip()
                nombre_prev = re.sub(r'\s*l\d+\s*$', '', nombre_prev, flags=re.IGNORECASE).strip()
                if nombre_prev:
                    cuit_to_nombre_previo[cuit] = nombre_prev

    return cuit_to_nombre_previo, cuit_to_cuota_previo


def procesar(
    imp_bytes, deu_bytes, imp_sheet,
    es_usd=False,
    tolerance=None,
    max_row=500,
    cuota_override=None,
    comprobantes_cache=None,
    mep_rates=None,
    log_fn=None,
    solo_mes=None,
    forzar_col_mes=None,
):
    """
    Corre la imputación en modo simulación.

    solo_mes: tupla (year, month) opcional. Si se pasa, se procesan SOLO las
    transferencias fechadas en ese mes y se escribe en la columna de ese mes.
    Sirve para semanas con cambio de mes en el medio (correr una pasada por mes).
    Si es None, autodetecta el mes más frecuente (comportamiento histórico).

    forzar_col_mes: tupla (year, month) opcional. Fuerza la COLUMNA del mes a
    escribir, pero SIN filtrar filas por fecha (procesa todas). Útil para la
    2da pasada de una semana con cambio de mes: "todo lo que no se imputó al mes
    previo va al mes nuevo" (lógica de próxima cuota impaga). Ignorado si se pasa
    solo_mes.
    Retorna (results, pago_menos, pago_mas, ambiguous, sin_fila, usd_en_pesos, mes_info, sheets_cfg).

    usd_en_pesos: clientes de la hoja '$  USD fijo' que pagaron en pesos
    (CUIT no está en las hojas de pesos pero sí en USD fijo). Se convierte el
    monto por el dólar MEP venta del día (mep_rates: dict 'YYYY-MM-DD' -> venta)
    y se reporta para que el usuario los habilite uno por uno.
    """
    if tolerance is None:
        tolerance = 5 if es_usd else 3000
    if cuota_override is None:
        cuota_override = {}
    if comprobantes_cache is None:
        comprobantes_cache = {}
    if log_fn is None:
        log_fn = lambda msg: None

    sheets_base = SHEETS_BASE_USD if es_usd else SHEETS_BASE_PESOS

    # Solo data_only para leer — no cargamos el workbook editable acá
    wb_deu_data = openpyxl.load_workbook(io.BytesIO(deu_bytes), data_only=True)
    wb_imp = openpyxl.load_workbook(io.BytesIO(imp_bytes))

    if solo_mes:
        tx_year, tx_month = solo_mes
        log_fn(f'Mes forzado (solo_mes): {MESES_ES.get(tx_month, "?")} {tx_year}')
    elif forzar_col_mes:
        tx_year, tx_month = forzar_col_mes
        log_fn(f'Columna forzada (forzar_col_mes): {MESES_ES.get(tx_month, "?")} {tx_year}')
    else:
        tx_year, tx_month = detectar_mes_transferencias(wb_imp[imp_sheet], max_row)
        if tx_year:
            log_fn(f'Mes detectado en transferencias: {MESES_ES.get(tx_month, "?")} {tx_year}')

    sheets_cfg, mes_info = build_sheets_cfg(wb_deu_data, sheets_base, year=tx_year, month=tx_month)
    cuit_index, nombre_index, cuota_history_cols = build_indices(wb_deu_data, sheets_cfg)
    log_fn(f'{len(cuit_index)} CUITs indexados en deudores')

    # En corridas de pesos, indexar también la hoja USD fijo para detectar
    # clientes USD que pagan su cuota en pesos (conversión por dólar MEP)
    usd_cuit_index, usd_hist_cols, usd_cfgs = {}, {}, {}
    if not es_usd:
        usd_cfgs, _ = build_sheets_cfg(wb_deu_data, SHEETS_BASE_USD, year=tx_year, month=tx_month)
        if usd_cfgs:
            usd_cuit_index, _, usd_hist_cols = build_indices(wb_deu_data, usd_cfgs)
            log_fn(f'{len(usd_cuit_index)} CUITs indexados en USD fijo')

    cuit_to_nombre_previo, cuit_to_cuota_previo = build_previo(wb_imp, imp_sheet, es_usd)
    log_fn(f'{len(cuit_to_nombre_previo)} CUITs con nombre desde hojas anteriores')
    log_fn(f'{len(comprobantes_cache)} CUITs en cache de comprobantes')

    def _buscar_nombre(nombre_str):
        return buscar_en_deudores_por_nombre(nombre_str, nombre_index)

    ws_imp = wb_imp[imp_sheet]
    results    = []
    ambiguous  = []
    pago_menos = []
    pago_mas   = []
    sin_fila   = []  # nombre conocido pero sin fila en deudores → prellena col H sin amarillo
    usd_en_pesos = []  # clientes USD fijo que pagaron en pesos (conversión MEP)
    written_deu_rows = {}  # (sname, srow) -> {'cuit': str, 'last_cuota': int}

    EXCESO_LIMITE   = 50 if es_usd else 50_000   # exceso máximo para imputar normalmente
    MULTI_TOL_RATIO = 0.05                        # tolerancia proporcional para múltiplos

    for row in ws_imp.iter_rows(min_row=4, max_row=max_row):
        fecha_val = row[0].value
        monto_val = row[5].value
        concepto = row[2].value
        col_h_val = row[7].value

        if fecha_val is None and monto_val is None:
            continue

        row_num = row[0].row

        if is_row_yellow(ws_imp, row_num):
            continue

        # Semana con cambio de mes: procesar solo las filas del mes pedido.
        if solo_mes:
            d_row = parse_date(fecha_val)
            if not (d_row and d_row.year == solo_mes[0] and d_row.month == solo_mes[1]):
                continue

        if col_h_val and ('PAGO MENOS' in str(col_h_val) or 'Saldo Disponible' in str(col_h_val)):
            continue

        cuit_raw = extract_cuit_from_concepto(concepto)
        if not cuit_raw:
            ambiguous.append({'row': row_num, 'motivo': 'Sin CUIT extraíble', 'concepto': str(concepto)[:60] if concepto else '', 'monto': monto_val, 'fecha': fecha_val})
            continue

        matches = cuit_index.get(cuit_raw, [])

        # Cliente de USD fijo pagando en pesos: el CUIT no está en las hojas de
        # pesos pero sí en la hoja USD. Convertir por MEP y reportar aparte.
        if not matches and cuit_raw in usd_cuit_index:
            umatches = usd_cuit_index[cuit_raw]
            # elegir la primera fila sin imputar este mes (real y cuota vacíos)
            destino = None
            for (u_sname, u_srow, u_snombre) in umatches:
                u_cfg = usd_cfgs[u_sname]
                u_real = wb_deu_data[u_sname].cell(u_srow, u_cfg['real_col']).value
                u_cuota_act = wb_deu_data[u_sname].cell(u_srow, u_cfg['cuota_col']).value
                ya_usada = (u_sname, u_srow) in written_deu_rows
                if u_real is None and not isinstance(u_cuota_act, (int, float)) and not ya_usada:
                    destino = (u_sname, u_srow, u_snombre, u_cuota_act)
                    break
            if destino is None:
                ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} en USD fijo pero el mes ya está imputado', 'cliente': umatches[0][2], 'cuit': cuit_raw, 'monto': monto_val})
                continue

            u_sname, u_srow, u_snombre, u_cuota_act = destino
            u_cfg = usd_cfgs[u_sname]
            fecha_dt = parse_date(fecha_val)
            tasa, tasa_fecha = mep_para_fecha(mep_rates or {}, fecha_dt) if fecha_dt else (None, None)

            monto_num = monto_val if isinstance(monto_val, (int, float)) else 0
            teo_usd = wb_deu_data[u_sname].cell(u_srow, u_cfg['teo_col']).value
            teo_usd = teo_usd if isinstance(teo_usd, (int, float)) else None
            equiv = round(monto_num / tasa, 2) if (tasa and monto_num) else None
            dif = round(equiv - teo_usd, 2) if (equiv is not None and teo_usd is not None) else None

            # número de cuota: misma lógica que el flujo normal, sobre la fila USD
            if isinstance(u_cuota_act, str) and 'parte' in u_cuota_act.lower():
                m = re.search(r'\d+', u_cuota_act)
                next_cuota = int(m.group()) + 1 if m else None
            else:
                max_hist = None
                for hc in usd_hist_cols.get(u_sname, []):
                    n_celda = max_cuota_celda(wb_deu_data[u_sname].cell(u_srow, hc).value)
                    if n_celda is not None and (max_hist is None or n_celda > max_hist):
                        max_hist = n_celda
                next_cuota = max_hist + 1 if max_hist is not None else None

            if next_cuota is None:
                ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} en USD fijo (pagó en pesos) pero no se pudo determinar cuota', 'cliente': u_snombre, 'cuit': cuit_raw, 'monto': monto_val, 'hoja': u_sname, 'hoja_fila': u_srow})
                continue

            usd_en_pesos.append({
                'imp_row': row_num,
                'cuit': cuit_raw,
                'cliente': u_snombre,
                'hoja': u_sname,
                'hoja_fila': u_srow,
                'fecha': fecha_dt,
                'monto_pesos': monto_num,
                'mep': tasa,
                'mep_fecha': tasa_fecha,
                'equiv_usd': equiv,
                'teo_usd': teo_usd,
                'dif_usd': dif,
                'cuota': next_cuota,
            })
            written_deu_rows[(u_sname, u_srow)] = {'cuit': cuit_raw, 'last_cuota': next_cuota}
            continue

        if not matches:
            nombre_previo = cuit_to_nombre_previo.get(cuit_raw)
            if nombre_previo:
                m2 = _buscar_nombre(nombre_previo)
                if len(m2) == 1:
                    matches = m2
                    log_fn(f'Fila {row_num}: fallback nombre "{nombre_previo}" → {m2[0][2]}')
                elif len(m2) > 1:
                    ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} no en deudores; nombre "{nombre_previo}" da {len(m2)} candidatos', 'concepto': str(concepto)[:60], 'monto': monto_val, 'fecha': fecha_val, 'matches': [(n, None) for _, _, n in m2]})
                    continue
                else:
                    sin_fila.append({'imp_row': row_num, 'nombre': nombre_previo})
                    ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} no en deudores; "{nombre_previo}" no encontrado en deudores', 'concepto': str(concepto)[:60], 'monto': monto_val, 'fecha': fecha_val})
                    continue
            else:
                nombre_comp = comprobantes_cache.get(cuit_raw)
                if nombre_comp:
                    m2 = _buscar_nombre(nombre_comp)
                    if len(m2) == 1:
                        matches = m2
                        log_fn(f'Fila {row_num}: fallback comprobantes "{nombre_comp}" → {m2[0][2]}')
                    elif len(m2) > 1:
                        ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} en comprobantes como "{nombre_comp}"; da {len(m2)} candidatos', 'concepto': str(concepto)[:60], 'monto': monto_val, 'fecha': fecha_val, 'matches': [(n, None) for _, _, n in m2]})
                        continue
                    else:
                        sin_fila.append({'imp_row': row_num, 'nombre': nombre_comp})
                        ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} en comprobantes como "{nombre_comp}"; no encontrado en deudores', 'concepto': str(concepto)[:60], 'monto': monto_val, 'fecha': fecha_val})
                        continue
                else:
                    ambiguous.append({'row': row_num, 'motivo': f'CUIT {cuit_raw} no encontrado en deudores, semanas anteriores ni comprobantes', 'concepto': str(concepto)[:60], 'monto': monto_val, 'fecha': fecha_val})
                    continue

        # monto numérico para todas las decisiones de abajo
        monto_num = monto_val if isinstance(monto_val, (int, float)) else 0

        # reparto_lotes=True → una cuota a cada lote (cada uno con su teórico)
        reparto_lotes = False
        if len(matches) > 1:
            candidatos = []
            for (sname, srow, snombre) in matches:
                cfg = sheets_cfg[sname]
                ws_d = wb_deu_data[sname]
                teo = ws_d.cell(srow, cfg['teo_col']).value
                real = ws_d.cell(srow, cfg['real_col']).value
                candidatos.append((sname, srow, snombre, teo, real))

            sin_imputar = [(s, r, n, t, rv) for s, r, n, t, rv in candidatos if rv is None]
            if not sin_imputar:
                ambiguous.append({'row': row_num, 'motivo': 'Todos los matches ya tienen el mes imputado', 'cuit': cuit_raw, 'monto': monto_val, 'matches': [(n, t) for _, _, n, t, _ in candidatos]})
                continue

            disponibles = [x for x in sin_imputar if (x[0], x[1]) not in written_deu_rows]
            if not disponibles:
                ambiguous.append({'row': row_num, 'motivo': 'Todos los lotes ya asignados en este run', 'cuit': cuit_raw, 'monto': monto_val, 'matches': [(n, t) for _, _, n, t, _ in sin_imputar]})
                continue

            if len(disponibles) > 1:
                # ¿El monto cubre la SUMA de los teóricos de los lotes sin imputar?
                # → repartir una cuota a cada lote (cada uno con su propio teórico).
                # No exige teóricos iguales (antes solo repartía si "empataban").
                suma_teo = sum((t or 0) for _, _, _, t, _ in disponibles)
                multi_tol = max(tolerance, suma_teo * MULTI_TOL_RATIO)
                if suma_teo > 0 and abs(monto_num - suma_teo) <= multi_tol:
                    targets = sorted(disponibles, key=lambda x: (x[0], x[1]))
                    reparto_lotes = True
                else:
                    # No alcanza para todos los lotes (pagó una sola cuota o un
                    # monto raro): imputar al lote cuyo teórico mejor coincide.
                    targets = [min(disponibles, key=lambda x: abs((x[3] or 0) - monto_num))]
            else:
                targets = [disponibles[0]]
        else:
            sname, srow, snombre = matches[0]
            cfg = sheets_cfg[sname]
            ws_d = wb_deu_data[sname]
            teo = ws_d.cell(srow, cfg['teo_col']).value
            real = ws_d.cell(srow, cfg['real_col']).value
            targets = [(sname, srow, snombre, teo, real)]

        sname, srow, snombre, teo_val, real_existente = targets[0]
        cfg = sheets_cfg[sname]

        if real_existente is not None:
            ambiguous.append({'row': row_num, 'motivo': f'Mes ya imputado ({real_existente}) en {sname} fila {srow}', 'cliente': snombre, 'cuit': cuit_raw, 'monto': monto_val})
            continue

        deu_key = (sname, srow)
        prev_assignment = written_deu_rows.get(deu_key)
        if prev_assignment is not None and prev_assignment['cuit'] != cuit_raw:
            ambiguous.append({'row': row_num, 'motivo': f'Destino duplicado (distinto CUIT) en {sname} fila {srow}', 'cliente': snombre, 'cuit': cuit_raw, 'monto': monto_val})
            continue

        teo_num = teo_val if isinstance(teo_val, (int, float)) else 0

        if reparto_lotes:
            # Una cuota a cada lote; cada lote lleva su propio teórico como pago real.
            usar = targets
            counts = [1] * len(usar)
            montos_por_lote = [t if isinstance(t, (int, float)) else 0
                               for _, _, _, t, _ in usar]
        else:
            if monto_num < teo_num and (teo_num - monto_num) > tolerance:
                pago_menos.append({'row': row_num, 'cliente': snombre, 'cuit': cuit_raw, 'transferido': monto_num, 'teorico': round(teo_num, 2 if es_usd else 0), 'diferencia': round(teo_num - monto_num, 2 if es_usd else 0)})
                continue

            # Exceso positivo: verificar si es múltiplo del teórico o excede el límite.
            # (Solo aplica a UN lote: el reparto entre lotes ya se decidió arriba.)
            n_cuotas = 1
            if teo_num > 0 and monto_num > teo_num + tolerance:
                exceso = monto_num - teo_num
                n = round(monto_num / teo_num)
                multi_tol = max(tolerance, teo_num * MULTI_TOL_RATIO)
                if n >= 2 and abs(monto_num - n * teo_num) <= multi_tol:
                    n_cuotas = n
                elif exceso > EXCESO_LIMITE:
                    pago_mas.append({'row': row_num, 'cliente': snombre, 'cuit': cuit_raw, 'transferido': monto_num, 'teorico': round(teo_num, 2 if es_usd else 0), 'diferencia': round(exceso, 2 if es_usd else 0)})
                    continue

            usar = [targets[0]]
            counts = [n_cuotas]
            monto_por_cuota = round(monto_num / n_cuotas, 2 if es_usd else 0)
            montos_por_lote = [monto_por_cuota]

        # Determinar la cuota de cada fila destino sin tocar estado, para poder
        # abortar limpio si alguna falla
        planes = []
        error_motivo = None
        for (p_sname, p_srow, p_snombre, p_teo, _p_real), cnt, monto_lote in zip(usar, counts, montos_por_lote):
            p_cfg = sheets_cfg[p_sname]
            prev = written_deu_rows.get((p_sname, p_srow))
            cuota_col_val = wb_deu_data[p_sname].cell(p_srow, p_cfg['cuota_col']).value

            if prev is not None:
                # Mismo cliente, misma fila deudores → cuota siguiente
                next_cuota = prev['last_cuota'] + 1
            elif isinstance(cuota_col_val, str) and 'parte' in cuota_col_val.lower():
                m = re.search(r'\d+', cuota_col_val)
                if m:
                    next_cuota = int(m.group()) + 1
                else:
                    error_motivo = f'Cuota dice "parte de..." pero no se pudo extraer número ({cuota_col_val!r})'
                    break
            else:
                max_hist = None
                for hc in cuota_history_cols.get(p_sname, []):
                    n_celda = max_cuota_celda(wb_deu_data[p_sname].cell(p_srow, hc).value)
                    if n_celda is not None and (max_hist is None or n_celda > max_hist):
                        max_hist = n_celda

                if max_hist is not None:
                    next_cuota = max_hist + 1
                elif cuit_raw in cuit_to_cuota_previo:
                    next_cuota = cuit_to_cuota_previo[cuit_raw] + 1
                    log_fn(f'Fila {row_num}: cuota previa {cuit_to_cuota_previo[cuit_raw]}+1={next_cuota}')
                elif cuit_raw in cuota_override:
                    next_cuota = cuota_override[cuit_raw]
                    log_fn(f'Fila {row_num}: cuota override {next_cuota} para {p_snombre}')
                else:
                    error_motivo = 'No se pudo determinar número de cuota (sin historial)'
                    break
            planes.append((p_sname, p_srow, p_snombre, p_teo, next_cuota, cnt, monto_lote))

        if error_motivo:
            ambiguous.append({'row': row_num, 'motivo': error_motivo, 'cliente': snombre, 'cuit': cuit_raw, 'hoja': sname, 'hoja_fila': srow})
            continue

        fecha_dt = parse_date(fecha_val)

        todos_matches = cuit_index.get(cuit_raw, []) or _buscar_nombre(cuit_to_nombre_previo.get(cuit_raw, ''))
        nombres_distintos = {str(n).strip().upper() for _, _, n in todos_matches}
        usar_lote = len(todos_matches) > 1 and len(nombres_distintos) == 1

        for p_sname, p_srow, p_snombre, p_teo, next_cuota, cnt, monto_lote in planes:
            p_cfg = sheets_cfg[p_sname]
            p_teo_num = p_teo if isinstance(p_teo, (int, float)) else 0
            if usar_lote:
                lote_val = wb_deu_data[p_sname].cell(p_srow, p_cfg['lote_col']).value
                lote_str = f' l{lote_val}' if lote_val is not None else ''
            else:
                lote_str = ''
            written_deu_rows[(p_sname, p_srow)] = {'cuit': cuit_raw, 'last_cuota': next_cuota + cnt - 1}
            for i in range(cnt):
                results.append({
                    'imp_row': row_num,
                    'cuit': cuit_raw,
                    'cliente': p_snombre,
                    'lote_str': lote_str,
                    'hoja': p_sname,
                    'hoja_fila': p_srow,
                    'monto_real': monto_lote,
                    'monto_teo': round(p_teo_num, 2 if es_usd else 0),
                    'diferencia': round(monto_lote - p_teo_num, 2 if es_usd else 0),
                    'cuota': next_cuota + i,
                    'fecha': fecha_dt,
                })

    wb_deu_data.close()
    wb_imp.close()

    # incluir la cfg de USD fijo para que aplicar() pueda escribir ahí
    sheets_cfg_out = dict(sheets_cfg)
    sheets_cfg_out.update(usd_cfgs)

    return results, pago_menos, pago_mas, ambiguous, sin_fila, usd_en_pesos, mes_info, sheets_cfg_out


def _format_cuotas(cuotas):
    if len(cuotas) == 1:
        return cuotas[0]
    if len(cuotas) == 2:
        return f'{cuotas[0]} y {cuotas[1]}'
    return ', '.join(str(c) for c in cuotas[:-1]) + f' y {cuotas[-1]}'


def _set_cell(cell, val):
    """Escribe val preservando el number_format de la celda. Excepción: si se
    escribe una fecha en una celda 'General' (típico de una columna de un mes
    nuevo aún sin formatear), se aplica formato de fecha para que no se vea como
    número de serie."""
    fmt = cell.number_format
    cell.value = val
    if isinstance(val, (datetime.datetime, datetime.date)) and fmt in (None, 'General'):
        fmt = 'dd/mm/yyyy'
    cell.number_format = fmt


def aplicar(results, pago_menos, imp_bytes, deu_bytes, imp_sheet, sheets_cfg, pago_mas=None, sin_fila=None, usd_en_pesos=None):
    """Carga los workbooks desde bytes, escribe y retorna (imp_bytes, deu_bytes).

    usd_en_pesos: solo las entradas que el usuario habilitó. Escribe en la hoja
    USD fijo el monto EN PESOS en "pago real", más cuota y fecha.
    """
    from collections import defaultdict
    wb_imp = openpyxl.load_workbook(io.BytesIO(imp_bytes))
    wb_deu_edit = openpyxl.load_workbook(io.BytesIO(deu_bytes))

    ws_edit = wb_imp[imp_sheet]

    for p in pago_menos:
        ws_edit.cell(p['row'], 8).value = 'PAGO MENOS'

    for p in (pago_mas or []):
        ws_edit.cell(p['row'], 8).value = 'PAGO MAS'

    # Prellenar nombre en col H sin amarillo (nombre conocido, fila en deudores no encontrada)
    for s in (sin_fila or []):
        cell = ws_edit.cell(s['imp_row'], 8)
        if not cell.value:  # no pisar si ya tiene algo
            cell.value = s['nombre']

    # Imputaciones: agrupar por imp_row (una transferencia puede cubrir N cuotas)
    imp_groups = defaultdict(list)
    for r in results:
        imp_groups[r['imp_row']].append(r)

    for row_num, grupo in imp_groups.items():
        filas = defaultdict(list)
        for r in grupo:
            filas[(r['hoja'], r['hoja_fila'])].append(r)
        if len(filas) == 1:
            cuotas = [r['cuota'] for r in grupo]
            nombre_corto = str(grupo[0]['cliente'])[:38]
            label = f"{nombre_corto}{grupo[0]['lote_str']} c{_format_cuotas(cuotas)}"
        else:
            # Transferencia repartida en varios lotes: "Nombre l24 c16 y l25 c16"
            segs = []
            for key in sorted(filas):
                rs = filas[key]
                cuotas_fila = _format_cuotas([r['cuota'] for r in rs])
                if rs[0]['lote_str']:
                    segs.append(f"{rs[0]['lote_str'].strip()} c{cuotas_fila}")
                else:
                    segs.append(f"{str(rs[0]['cliente'])[:25]} c{cuotas_fila}")
            if grupo[0]['lote_str']:
                label = f"{str(grupo[0]['cliente'])[:38]} " + ' y '.join(segs)
            else:
                label = ' y '.join(segs)
        ws_edit.cell(row_num, 8).value = label
        for cell in ws_edit[row_num]:
            cell.fill = YELLOW_FILL

    # Clientes USD fijo que pagaron en pesos (solo los habilitados)
    for e in (usd_en_pesos or []):
        ws_edit.cell(e['imp_row'], 8).value = f"{str(e['cliente'])[:38]} c{e['cuota']}"
        for cell in ws_edit[e['imp_row']]:
            cell.fill = YELLOW_FILL
        cfg = sheets_cfg[e['hoja']]
        ws_deu = wb_deu_edit[e['hoja']]
        for col, val in [
            (cfg['real_col'],  e['monto_pesos']),
            (cfg['cuota_col'], e['cuota']),
            (cfg['fecha_col'], e['fecha']),
        ]:
            _set_cell(ws_deu.cell(e['hoja_fila'], col), val)

    # Deudores: agrupar por fila (mismo lote puede tener varias cuotas)
    deu_groups = defaultdict(list)
    for r in results:
        deu_groups[(r['hoja'], r['hoja_fila'])].append(r)

    for (sname, srow), grupo in deu_groups.items():
        cfg = sheets_cfg[sname]
        ws_deu = wb_deu_edit[sname]
        cuotas = [r['cuota'] for r in grupo]
        for col, val in [
            (cfg['real_col'],  sum(r['monto_real'] for r in grupo)),
            (cfg['cuota_col'], _format_cuotas(cuotas)),
            (cfg['fecha_col'], grupo[0]['fecha']),
        ]:
            _set_cell(ws_deu.cell(srow, col), val)

    imp_out = io.BytesIO()
    deu_out = io.BytesIO()
    wb_imp.save(imp_out)
    wb_deu_edit.save(deu_out)
    wb_imp.close()
    wb_deu_edit.close()
    return imp_out.getvalue(), deu_out.getvalue()
