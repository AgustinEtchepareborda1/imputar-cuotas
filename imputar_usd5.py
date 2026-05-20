"""
Imputation script for USD 5 -> deudores (hoja '$ USD fijo').
Run with DRY_RUN=True first to review before writing.
"""

import openpyxl
from openpyxl.styles import PatternFill
import re
import datetime
import shutil
import os
from comprobantes_helper import cargar_indice as _cargar_comprobantes

# ── CONFIG ──────────────────────────────────────────────────────────────────
DRY_RUN = True   # ← cambiar a False para escribir realmente

IMP_FILE  = r'datos/IMPUTACIONES -20.05.26.xlsx'
DEU_FILE  = r'datos/DEUDORES FINK 20-05-2026.xlsx'
IMP_SHEET = 'USD 5'
REF_SHEET = 'USD 4'   # solo para referencia visual, no se toca

TOLERANCE = 5          # diferencia máxima aceptable en USD
MAX_ROW   = 9          # procesar solo hasta esta fila de imputaciones

# Overrides manuales de cuota cuando no se puede determinar del historial
CUOTA_OVERRIDE = {
    # '20XXXXXXXXX': 15,
}

YELLOW_FILL = PatternFill(patternType='solid', fgColor='FFFF00')

# ── Estructura de la hoja de deudores USD ────────────────────────────────────
# '$  USD fijo' (dos espacios): header en fila 3, datos desde fila 4
# Columnas Mayo 2026: teo=131, real=132, cuota=133, fecha=134
SHEETS_CFG = {
    '$  USD fijo': {
        'header_row': 3, 'data_start': 4,
        'cuit_col':   13,  # M  DNI o CUIT (formato "XX-XXXXXXXX-X", puede tener varios separados por /)
        'nombre_col': 10,  # J  Nombre y Apellido
        'lote_col':    9,  # I  LOTE
        'teo_col':   131,  # pago MAYO 26 teorico
        'real_col':  132,  # pago MAYO 26 real
        'cuota_col': 133,  # NUMERO DE CUOTA
        'fecha_col': 134,  # fecha depo
        'max_cuota_col': 130,  # MAYOR CUOTA (fórmula, no usada directamente)
    },
}

# ── Helpers ──────────────────────────────────────────────────────────────────

def normalize_cuits(raw):
    """Devuelve lista de strings de dígitos extraídos del campo CUIT."""
    if not raw:
        return []
    results = []
    for part in str(raw).split('/'):
        digits = re.sub(r'[^0-9]', '', part.strip())
        if len(digits) >= 10:   # CUITs válidos: 11 dígitos; también acepta 12
            results.append(digits)
    return results


def is_row_yellow(ws, row_num):
    """Devuelve True si la primera celda con valor tiene fondo amarillo."""
    for cell in ws[row_num]:
        if cell.value is None:
            continue
        fill = cell.fill
        if fill and fill.patternType and fill.patternType != 'none':
            rgb = str(fill.fgColor.rgb) if fill.fgColor else ''
            if 'FFFF00' in rgb or rgb == 'FFFFFF00':
                return True
        break
    return False


def extract_cuit_from_concepto(concepto):
    """
    Extrae números de 11 o 12 dígitos del campo Concepto de imputaciones.
    Devuelve el primer número encontrado, o None.
    """
    if not concepto:
        return None
    matches = re.findall(r'\b(\d{11,12})\b', str(concepto))
    return matches[0] if matches else None


def parse_date(val):
    """Convierte la fecha de imputaciones (string 'DD-MM-YYYY' o datetime) a datetime."""
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, str):
        for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


# ── 1. Cargar deudores ────────────────────────────────────────────────────────
print('Cargando deudores USD...')
wb_deu_data = openpyxl.load_workbook(DEU_FILE, data_only=True)
wb_deu_edit = openpyxl.load_workbook(DEU_FILE)

# ── 2. Construir índice CUIT → (sheet, row, nombre) ──────────────────────────
cuit_index = {}

for sheet_name, cfg in SHEETS_CFG.items():
    ws_data = wb_deu_data[sheet_name]
    max_row = ws_data.max_row
    for r in range(cfg['data_start'], max_row + 1):
        nombre = ws_data.cell(r, cfg['nombre_col']).value
        cuit_raw = ws_data.cell(r, cfg['cuit_col']).value
        if not nombre and not cuit_raw:
            continue
        cuits = normalize_cuits(cuit_raw)
        for c in cuits:
            cuit_index.setdefault(c, []).append((sheet_name, r, nombre))

print(f'  -> {len(cuit_index)} CUITs indexados en deudores USD')

# ── 2a. Pre-computar columnas históricas de "NUMERO DE CUOTA" ────────────────
cuota_history_cols = {}
for sheet_name, cfg in SHEETS_CFG.items():
    ws_d = wb_deu_data[sheet_name]
    cols = []
    for c in range(1, ws_d.max_column + 1):
        h = ws_d.cell(cfg['header_row'], c).value
        if h and ('numero' in str(h).lower() or 'n°' in str(h).lower() or 'nro' in str(h).lower()) and 'cuota' in str(h).lower():
            if c != cfg['cuota_col']:
                cols.append(c)
    cuota_history_cols[sheet_name] = cols
    print(f'  -> {sheet_name}: {len(cols)} columnas históricas de cuota')

# ── 2b. Índice nombre → deudores ─────────────────────────────────────────────
nombre_index = {}
for sheet_name, cfg in SHEETS_CFG.items():
    ws_data = wb_deu_data[sheet_name]
    for r in range(cfg['data_start'], ws_data.max_row + 1):
        nombre = ws_data.cell(r, cfg['nombre_col']).value
        if not nombre:
            continue
        for palabra in str(nombre).upper().split():
            if len(palabra) >= 4:
                nombre_index.setdefault(palabra, []).append((sheet_name, r, nombre))

# ── 2c. Índice CUIT → nombre desde hojas anteriores de imputaciones ──────────
# Solo considera hojas USD anteriores (USD 4, USD 3, USD 2, USD)
wb_imp_ref = openpyxl.load_workbook(IMP_FILE)
all_sheets = wb_imp_ref.sheetnames

try:
    idx_actual = all_sheets.index(IMP_SHEET)
    # Solo hojas USD anteriores
    hojas_previas = [h for h in reversed(all_sheets[:idx_actual]) if 'usd' in h.lower()]
except ValueError:
    hojas_previas = []

cuit_to_nombre_previo = {}
cuit_to_cuota_previo  = {}
for hoja in hojas_previas[:6]:
    try:
        ws_prev = wb_imp_ref[hoja]
    except Exception:
        continue
    for row in ws_prev.iter_rows(min_row=4, max_row=ws_prev.max_row):
        concepto = row[2].value if len(row) > 2 else None
        col_h    = row[7].value if len(row) > 7 else None
        if not concepto or not col_h:
            continue
        col_h_str = str(col_h).strip()
        if not col_h_str or 'PAGO MENOS' in col_h_str or 'Saldo' in col_h_str:
            continue
        cuit = extract_cuit_from_concepto(concepto)
        if cuit and cuit not in cuit_to_nombre_previo:
            m_cuota = re.search(r'\bc(\d+)\s*$', col_h_str, re.IGNORECASE)
            if m_cuota:
                cuit_to_cuota_previo[cuit] = int(m_cuota.group(1))
            nombre_prev = re.sub(r'\s+c\d+\s*$', '', col_h_str, flags=re.IGNORECASE).strip()
            if nombre_prev:
                cuit_to_nombre_previo[cuit] = nombre_prev

print(f'  -> {len(cuit_to_nombre_previo)} CUITs con nombre desde hojas USD anteriores')

# ── 2d. Cache de comprobantes del bot (Google Sheets) ────────────────────────
cuit_to_nombre_comprobantes = _cargar_comprobantes()
print(f'  -> {len(cuit_to_nombre_comprobantes)} CUITs en cache de comprobantes')


def buscar_en_deudores_por_nombre(nombre_str):
    """Busca en deudores por palabras del nombre. Devuelve lista de (sheet, row, nombre)."""
    palabras = [p.upper() for p in nombre_str.split() if len(p) >= 4]
    if not palabras:
        return []
    sets = [set((s, r) for s, r, _ in nombre_index.get(p, [])) for p in palabras]
    if not sets:
        return []
    comunes = sets[0]
    for s in sets[1:]:
        comunes &= s
    if not comunes:
        from collections import Counter
        cnt = Counter()
        for p in palabras:
            for s, r, _ in nombre_index.get(p, []):
                cnt[(s, r)] += 1
        max_hits = max(cnt.values()) if cnt else 0
        comunes = {k for k, v in cnt.items() if v == max_hits and v >= max(2, len(palabras) - 1)}
    resultado = []
    seen = set()
    for p in palabras:
        for s, r, n in nombre_index.get(p, []):
            if (s, r) in comunes and (s, r) not in seen:
                resultado.append((s, r, n))
                seen.add((s, r))
    return resultado


# ── 3. Leer hoja USD ─────────────────────────────────────────────────────────
print(f'\nCargando imputaciones ({IMP_SHEET})...')
wb_imp = openpyxl.load_workbook(IMP_FILE)
ws_imp = wb_imp[IMP_SHEET]

# ── 4. Procesar cada fila pendiente ──────────────────────────────────────────
results   = []
ambiguous = []
pago_menos= []
written_deu_rows = set()

for row in ws_imp.iter_rows(min_row=4, max_row=MAX_ROW):
    fecha_val  = row[0].value  # col A
    monto_val  = row[5].value  # col F (Crédito USD)
    concepto   = row[2].value  # col C
    col_h_val  = row[7].value  # col H (notas/nombre)

    if fecha_val is None and monto_val is None:
        continue

    row_num = row[0].row

    if is_row_yellow(ws_imp, row_num):
        continue

    if col_h_val and ('PAGO MENOS' in str(col_h_val) or 'Saldo Disponible' in str(col_h_val)):
        continue

    cuit_raw = extract_cuit_from_concepto(concepto)
    if not cuit_raw:
        ambiguous.append({
            'row': row_num,
            'motivo': 'Sin CUIT extraíble',
            'concepto': concepto,
            'monto': monto_val,
            'fecha': fecha_val,
        })
        continue

    matches = cuit_index.get(cuit_raw, [])
    if not matches:
        nombre_previo = cuit_to_nombre_previo.get(cuit_raw)
        if nombre_previo:
            matches_nombre = buscar_en_deudores_por_nombre(nombre_previo)
            if len(matches_nombre) == 1:
                matches = matches_nombre
                print(f'    [fallback nombre] Fila {row_num}: CUIT {cuit_raw} -> "{nombre_previo}" -> {matches[0][2]}')
            elif len(matches_nombre) > 1:
                ambiguous.append({
                    'row': row_num,
                    'motivo': f'CUIT {cuit_raw} no en deudores; nombre "{nombre_previo}" da {len(matches_nombre)} candidatos',
                    'concepto': concepto,
                    'monto': monto_val,
                    'fecha': fecha_val,
                    'matches': [(n, None) for _, _, n in matches_nombre],
                })
                continue
            else:
                ambiguous.append({
                    'row': row_num,
                    'motivo': f'CUIT {cuit_raw} no en deudores; nombre "{nombre_previo}" no encontrado en deudores',
                    'concepto': concepto,
                    'monto': monto_val,
                    'fecha': fecha_val,
                })
                continue
        else:
            # Fallback: buscar en cache de comprobantes del bot
            nombre_comprobante = cuit_to_nombre_comprobantes.get(cuit_raw)
            if nombre_comprobante:
                matches_nombre = buscar_en_deudores_por_nombre(nombre_comprobante)
                if len(matches_nombre) == 1:
                    matches = matches_nombre
                    print(f'    [fallback comprobantes] Fila {row_num}: CUIT {cuit_raw} -> "{nombre_comprobante}" -> {matches[0][2]}')
                elif len(matches_nombre) > 1:
                    ambiguous.append({
                        'row': row_num,
                        'motivo': f'CUIT {cuit_raw} en comprobantes como "{nombre_comprobante}"; da {len(matches_nombre)} candidatos en deudores',
                        'concepto': concepto,
                        'monto': monto_val,
                        'fecha': fecha_val,
                        'matches': [(n, None) for _, _, n in matches_nombre],
                    })
                    continue
                else:
                    ambiguous.append({
                        'row': row_num,
                        'motivo': f'CUIT {cuit_raw} en comprobantes como "{nombre_comprobante}"; no encontrado en deudores',
                        'concepto': concepto,
                        'monto': monto_val,
                        'fecha': fecha_val,
                    })
                    continue
            else:
                ambiguous.append({
                    'row': row_num,
                    'motivo': f'CUIT {cuit_raw} no encontrado en deudores USD, semanas anteriores ni comprobantes',
                    'concepto': concepto,
                    'monto': monto_val,
                    'fecha': fecha_val,
                })
                continue

    if len(matches) > 1:
        candidatos = []
        for (sname, srow, snombre) in matches:
            cfg = SHEETS_CFG[sname]
            ws_d = wb_deu_data[sname]
            teo  = ws_d.cell(srow, cfg['teo_col']).value
            real = ws_d.cell(srow, cfg['real_col']).value
            candidatos.append((sname, srow, snombre, teo, real))

        sin_imputar = [(s,r,n,t,rv) for s,r,n,t,rv in candidatos if rv is None]
        if len(sin_imputar) == 0:
            ambiguous.append({
                'row': row_num,
                'motivo': 'Todos los matches ya tienen Mayo 26 imputado',
                'cuit': cuit_raw,
                'monto': monto_val,
                'matches': [(n,t) for _,_,n,t,_ in candidatos],
            })
            continue
        if len(sin_imputar) > 1:
            mejor = min(sin_imputar, key=lambda x: abs((x[3] or 0) - (monto_val or 0)))
            diferencia_mejor = abs((mejor[3] or 0) - (monto_val or 0))
            otros_con_mismo_monto = [
                x for x in sin_imputar
                if x != mejor and abs(abs((x[3] or 0) - (monto_val or 0)) - diferencia_mejor) < 0.01
            ]
            if otros_con_mismo_monto:
                disponibles = [x for x in sin_imputar if (x[0], x[1]) not in written_deu_rows]
                if not disponibles:
                    ambiguous.append({
                        'row': row_num,
                        'motivo': 'Todos los lotes con mismo CUIT ya asignados en este run',
                        'cuit': cuit_raw,
                        'monto': monto_val,
                        'matches': [(n, t) for _,_,n,t,_ in sin_imputar],
                    })
                    continue
                selected = sorted(disponibles, key=lambda x: (x[0], x[1]))[0]
            else:
                selected = mejor
        else:
            selected = sin_imputar[0]
    else:
        sname, srow, snombre = matches[0]
        cfg   = SHEETS_CFG[sname]
        ws_d  = wb_deu_data[sname]
        teo   = ws_d.cell(srow, cfg['teo_col']).value
        real  = ws_d.cell(srow, cfg['real_col']).value
        selected = (sname, srow, snombre, teo, real)

    sname, srow, snombre, teo_val, real_existente = selected
    cfg = SHEETS_CFG[sname]

    if real_existente is not None or (sname, srow) in written_deu_rows:
        ambiguous.append({
            'row': row_num,
            'motivo': f'Mayo 26 ya imputado ({real_existente}) o destino duplicado en {sname} fila {srow}',
            'cliente': snombre,
            'cuit': cuit_raw,
            'monto': monto_val,
        })
        continue

    monto_num = monto_val if isinstance(monto_val, (int, float)) else 0
    teo_num   = teo_val   if isinstance(teo_val,  (int, float)) else 0

    if monto_num < teo_num and (teo_num - monto_num) > TOLERANCE:
        pago_menos.append({
            'row': row_num,
            'cliente': snombre,
            'cuit': cuit_raw,
            'transferido': monto_num,
            'teorico': round(teo_num, 2),
            'diferencia': round(teo_num - monto_num, 2),
        })
        continue

    # Determinar número de cuota
    ws_d = wb_deu_data[sname]
    cuota_col_val = ws_d.cell(srow, cfg['cuota_col']).value

    if isinstance(cuota_col_val, str) and 'parte' in cuota_col_val.lower():
        m = re.search(r'\d+', cuota_col_val)
        if m:
            next_cuota = int(m.group()) + 1
        else:
            ambiguous.append({
                'row': row_num,
                'motivo': f'Cuota dice "parte de..." pero no se pudo extraer número ({cuota_col_val!r})',
                'cliente': snombre,
                'cuit': cuit_raw,
                'hoja': sname,
                'hoja_fila': srow,
            })
            continue
    else:
        hist_cols = cuota_history_cols.get(sname, [])
        max_hist = None
        for hc in hist_cols:
            val = ws_d.cell(srow, hc).value
            if isinstance(val, (int, float)) and 0 < val <= 200:
                if max_hist is None or val > max_hist:
                    max_hist = int(val)

        if max_hist is not None:
            next_cuota = max_hist + 1
        elif cuit_raw in cuit_to_cuota_previo:
            next_cuota = cuit_to_cuota_previo[cuit_raw] + 1
            print(f'    [cuota imputaciones] Fila {row_num}: usando cuota previa {cuit_to_cuota_previo[cuit_raw]}+1={next_cuota}')
        elif cuit_raw in CUOTA_OVERRIDE:
            next_cuota = CUOTA_OVERRIDE[cuit_raw]
            print(f'    [cuota override] Fila {row_num}: usando cuota manual {next_cuota} para {snombre}')
        else:
            ambiguous.append({
                'row': row_num,
                'motivo': 'No se pudo determinar número de cuota (sin historial en deudores ni en imputaciones)',
                'cliente': snombre,
                'cuit': cuit_raw,
                'hoja': sname,
                'hoja_fila': srow,
            })
            continue

    fecha_dt = parse_date(fecha_val)

    todos_matches = cuit_index.get(cuit_raw, []) or buscar_en_deudores_por_nombre(
        cuit_to_nombre_previo.get(cuit_raw, ''))
    nombres_distintos = {str(n).strip().upper() for _, _, n in todos_matches}
    if len(todos_matches) > 1 and len(nombres_distintos) == 1:
        lote_val = wb_deu_data[sname].cell(srow, cfg['lote_col']).value
        lote_str = f' l{lote_val}' if lote_val is not None else ''
    else:
        lote_str = ''

    written_deu_rows.add((sname, srow))
    results.append({
        'imp_row': row_num,
        'cuit': cuit_raw,
        'cliente': snombre,
        'lote_str': lote_str,
        'hoja': sname,
        'hoja_fila': srow,
        'monto_real': monto_num,
        'monto_teo': round(teo_num, 2),
        'diferencia': round(monto_num - teo_num, 2),
        'cuota': next_cuota,
        'fecha': fecha_dt,
    })

# ── 5. Reporte ────────────────────────────────────────────────────────────────
print('\n' + '='*70)
print(f'  RESULTADOS DRY-RUN  (DRY_RUN={DRY_RUN})')
print('='*70)

print(f'\nOK PARA IMPUTAR ({len(results)}):')
for r in results:
    dif_str = f'+{r["diferencia"]}' if r['diferencia'] >= 0 else str(r['diferencia'])
    lote_disp = r.get('lote_str', '')
    print(f'  Fila {r["imp_row"]:3d}: {r["cliente"][:35]:<35} | '
          f'CUIT {r["cuit"]} | '
          f'U$D {r["monto_real"]:>8.2f} (teo {r["monto_teo"]:>8.2f}, dif {dif_str:>8}) | '
          f'{lote_disp}c{r["cuota"]} | {r["fecha"].strftime("%d/%m/%Y") if r["fecha"] else "?"} | '
          f'{r["hoja"]} fila {r["hoja_fila"]}')

print(f'\nPAGO MENOS ({len(pago_menos)}) -- se escribira "PAGO MENOS" en col H:')
for p in pago_menos:
    print(f'  Fila {p["row"]:3d}: {p["cliente"][:35]:<35} | '
          f'CUIT {p["cuit"]} | '
          f'U$D {p["transferido"]:>8.2f} vs teo {p["teorico"]:>8.2f} | '
          f'diferencia {p["diferencia"]:>8.2f}')

print(f'\nCASOS AMBIGUOS ({len(ambiguous)}) -- sin imputar, revisar manualmente:')
for a in ambiguous:
    motivo = a.get('motivo', '')
    fila   = a.get('row', '?')
    cliente= a.get('cliente', a.get('concepto', ''))
    monto  = a.get('monto', '')
    matches= a.get('matches', '')
    print(f'  Fila {fila:3d}: {motivo}')
    if cliente:
        print(f'         Cliente/Concepto: {str(cliente)[:60]}')
    if monto:
        print(f'         Monto: U$D {monto:>8.2f}' if isinstance(monto, (int,float)) else f'         Monto: {monto}')
    if matches:
        print(f'         Candidatos: {matches}')

print('\n' + '='*70)
print(f'  Total filas {IMP_SHEET} a procesar: {len(results)+len(pago_menos)+len(ambiguous)}')
print(f'  -> Imputar: {len(results)} | PAGO MENOS: {len(pago_menos)} | Revisar: {len(ambiguous)}')
print('='*70)

# ── 6. Escritura (solo si DRY_RUN=False) ─────────────────────────────────────
if DRY_RUN:
    print('\n[DRY RUN] No se modifico ningun archivo.')
    print('  Cambia DRY_RUN = False y volvé a ejecutar para aplicar.')
else:
    for path in [IMP_FILE, DEU_FILE]:
        bak = path + '.bak'
        if not os.path.exists(bak):
            shutil.copy2(path, bak)
            print(f'Backup: {bak}')

    ws_imp_edit = wb_imp[IMP_SHEET]

    for p in pago_menos:
        row_num = p['row']
        ws_imp_edit.cell(row_num, 8).value = 'PAGO MENOS'

    for r in results:
        row_num  = r['imp_row']
        sname    = r['hoja']
        srow     = r['hoja_fila']
        cfg      = SHEETS_CFG[sname]
        ws_edit  = wb_deu_edit[sname]

        ws_edit.cell(srow, cfg['real_col']).value  = r['monto_real']
        ws_edit.cell(srow, cfg['cuota_col']).value = r['cuota']
        ws_edit.cell(srow, cfg['fecha_col']).value = r['fecha']

        nombre_corto = str(r['cliente'])[:38]
        ws_imp_edit.cell(row_num, 8).value = f"{nombre_corto}{r['lote_str']} c{r['cuota']}"
        for cell in ws_imp_edit[row_num]:
            cell.fill = YELLOW_FILL

    wb_deu_edit.save(DEU_FILE)
    wb_imp.save(IMP_FILE)
    print(f'\nArchivos guardados. Imputadas: {len(results)} | PAGO MENOS: {len(pago_menos)}')
