"""
Imputation script for S 120 -> deudores.
Run with DRY_RUN=True first to review before writing.
"""

import openpyxl
from openpyxl.styles import PatternFill
import re
import unicodedata
import datetime
import shutil
import os
from comprobantes_helper import cargar_indice as _cargar_comprobantes
from imputar_core import max_cuota_celda

# ── CONFIG ──────────────────────────────────────────────────────────────────
DRY_RUN = True   # ← cambiar a False para escribir realmente

IMP_FILE  = r'datos/IMPUTACIONES -20.05.26(1).xlsx'
DEU_FILE  = r'datos/DEUDORES FINK 20-05-2026(1).xlsx'
IMP_SHEET = None   # None = auto-detectar última hoja 'S NNN' sin procesar

TOLERANCE = 3000       # diferencia máxima aceptable en pesos
MAX_ROW   = None       # None = procesar hasta el final de la hoja

# Overrides manuales de cuota cuando max_cuota=None y el número correcto se conoce
# Formato: 'CUIT': cuota_a_imputar
CUOTA_OVERRIDE = {
    '20358346740': 12,   # Bustamante Gustavo (l6 y l7)
    # '20168468939': ??, # Castañeda Miguel Fabian — completar cuota
}

YELLOW_FILL = PatternFill(patternType='solid', fgColor='FFFF00')

# ── Estructura de cada hoja de deudores ─────────────────────────────────────
# header_row: fila donde están los títulos de columnas
# data_start: primera fila de datos
SHEETS_CFG = {
    'INDICE CAC': {
        'header_row': 5, 'data_start': 6,
        'cuit_col': 12,   # L  DNI o CUIT
        'nombre_col': 9,  # I  Nombre y Apellido
        'lote_col':  8,   # H  LOTE
        'max_cuota_col':1,# A  MAYOR CUOTA (fórmula)
        # teo_col / real_col / cuota_col / fecha_col → auto-detectados por mes
    },
    'INDICE CAC M. OBRA': {
        'header_row': 3, 'data_start': 4,
        'cuit_col': 11,   # K  DNI o CUIT
        'nombre_col': 9,  # I  Nombre y Apellido
        'lote_col':  8,   # H  LOTE
        'max_cuota_col':1,
    },
    'BOLSA CEMENTO': {
        'header_row': 3, 'data_start': 4,
        'cuit_col': 14,   # N  DNI o CUIT
        'nombre_col': 10, # J  Nombre y Apellido
        'lote_col':  9,   # I  LOTE
        'max_cuota_col':1,
    },
}

# ── Helpers ──────────────────────────────────────────────────────────────────

MESES_ES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
    5: 'mayo',  6: 'junio',   7: 'julio', 8: 'agosto',
    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre',
}


def detectar_columnas_mes(wb_deu, sheets_cfg, mes, anio):
    """Busca en cada hoja las columnas del mes/año dado y actualiza sheets_cfg."""
    mes_str   = MESES_ES[mes]
    anio_short = str(anio)[-2:]
    for sheet_name, cfg in sheets_cfg.items():
        ws = wb_deu[sheet_name]
        hrow = cfg['header_row']
        teo_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(hrow, c).value
            if not h:
                continue
            h_lower = str(h).lower()
            if mes_str in h_lower and anio_short in h_lower and 'teorico' in h_lower:
                teo_col = c
                break
        if teo_col is None:
            raise ValueError(
                f"No se encontró '{mes_str} {anio_short} teorico' en '{sheet_name}' "
                f"(fila {hrow}). ¿El archivo de deudores tiene ese mes?"
            )
        cfg['teo_col']   = teo_col
        cfg['real_col']  = teo_col + 1
        cfg['cuota_col'] = teo_col + 2
        cfg['fecha_col'] = teo_col + 3
        col_letra = ws.cell(hrow, teo_col).column_letter
        print(f'  -> {sheet_name}: col {col_letra} = {mes_str} {anio_short} teorico')


def normalize_str(s):
    """Convierte a mayúsculas y quita tildes/diacríticos para comparación."""
    s = str(s).upper()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def normalize_cuits(raw):
    """Extrae todos los CUITs de una celda: guiones, '/' o texto libre entre CUITs."""
    if raw is None or raw == '':
        return []
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = str(raw)
    # Unir grupos de dígitos conectados por guion/punto (con espacios sueltos
    # alrededor) mientras no excedan los 12 dígitos de un CUIT.
    results, cur, prev_end = [], '', None
    for g in re.finditer(r'\d+', s):
        sep = s[prev_end:g.start()] if prev_end is not None else None
        unido = sep is not None and re.fullmatch(r'\s*[.\-]\s*', sep)
        if cur and unido and len(cur) + len(g.group()) <= 12:
            cur += g.group()
        else:
            if 10 <= len(cur) <= 12:
                results.append(cur)
            cur = g.group()
        prev_end = g.end()
    if 10 <= len(cur) <= 12:
        results.append(cur)
    if not results:
        # formato raro (ej: dígitos separados solo por espacios): unir todo
        for part in s.split('/'):
            digits = re.sub(r'\D', '', part)
            if len(digits) >= 10:
                results.append(digits)
    return list(dict.fromkeys(results))


def is_row_yellow(ws, row_num):
    """Devuelve True si la primera celda con valor tiene fondo amarillo."""
    for cell in ws[row_num]:
        if cell.value is None:
            continue
        fill = cell.fill
        if fill and fill.patternType and fill.patternType != 'none':
            rgb = str(fill.fgColor.rgb) if fill.fgColor else ''
            if 'FFFF00' in rgb or rgb == 'FFFFFF00':
                return True
        break
    return False


def extract_cuit_from_concepto(concepto):
    """
    Extrae números de 11 o 12 dígitos del campo Concepto de imputaciones.
    Devuelve el primer número encontrado, o None.
    """
    if not concepto:
        return None
    matches = re.findall(r'\b(\d{11,12})\b', str(concepto))
    return matches[0] if matches else None


def parse_date(val):
    """Convierte la fecha de imputaciones (string 'DD-MM-YYYY' o datetime) a datetime."""
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, str):
        for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


# ── 0. Auto-detectar hoja S y mes de imputación ─────────────────────────────
from collections import Counter as _Counter

_wb_pre = openpyxl.load_workbook(IMP_FILE, read_only=True)
_all_sheets = _wb_pre.sheetnames

if IMP_SHEET is None:
    hojas_s = [s for s in _all_sheets if re.match(r'^S\s+\d+$', s, re.IGNORECASE)]
    if not hojas_s:
        raise ValueError('No se encontraron hojas "S NNN" en el archivo de imputaciones.')
    IMP_SHEET = hojas_s[-1]
    print(f'Hoja auto-detectada: {IMP_SHEET}')
else:
    print(f'Hoja: {IMP_SHEET}')

_ws_pre = _wb_pre[IMP_SHEET]
_fechas = [parse_date(r[0]) for r in _ws_pre.iter_rows(min_row=4, values_only=True) if r[0]]
_fechas = [f for f in _fechas if f]
_wb_pre.close()

if not _fechas:
    raise ValueError(f'No se encontraron fechas válidas en la hoja {IMP_SHEET}.')

_mes_imp, _anio_imp = _Counter((f.month, f.year) for f in _fechas).most_common(1)[0][0]
print(f'Mes detectado: {MESES_ES[_mes_imp].upper()} {_anio_imp}')

# ── 1. Cargar deudores (dos veces: una para valores, otra para escritura) ──
print('Cargando deudores...')
wb_deu_data = openpyxl.load_workbook(DEU_FILE, data_only=True)   # leer valores
wb_deu_edit = openpyxl.load_workbook(DEU_FILE)                    # para escritura

print('Detectando columnas del mes en deudores...')
detectar_columnas_mes(wb_deu_data, SHEETS_CFG, _mes_imp, _anio_imp)

# ── 2. Construir índice CUIT → (sheet, row, nombre, ws_edit) ─────────────
cuit_index = {}   # normalized_cuit_str -> list of (sheet_name, row, nombre)

for sheet_name, cfg in SHEETS_CFG.items():
    ws_data = wb_deu_data[sheet_name]
    max_row = ws_data.max_row
    for r in range(cfg['data_start'], max_row + 1):
        nombre = ws_data.cell(r, cfg['nombre_col']).value
        cuit_raw = ws_data.cell(r, cfg['cuit_col']).value
        if not nombre and not cuit_raw:
            continue
        cuits = normalize_cuits(cuit_raw)
        for c in cuits:
            cuit_index.setdefault(c, []).append((sheet_name, r, nombre))

print(f'  -> {len(cuit_index)} CUITs indexados en deudores')

# ── 2a. Pre-computar columnas históricas de "NUMERO DE CUOTA" por hoja ───────
# Sirven como fallback cuando la fórmula MAYOR CUOTA (col A) da None.
cuota_history_cols = {}  # sheet_name -> [col_idx, ...]
for sheet_name, cfg in SHEETS_CFG.items():
    ws_d = wb_deu_data[sheet_name]
    cols = []
    for c in range(1, ws_d.max_column + 1):
        h = ws_d.cell(cfg['header_row'], c).value
        if h and ('numero' in str(h).lower() or 'n°' in str(h).lower() or 'nro' in str(h).lower()) and 'cuota' in str(h).lower():
            if c != cfg['cuota_col']:  # excluir el mes actual
                cols.append(c)
    cuota_history_cols[sheet_name] = cols
    print(f'  -> {sheet_name}: {len(cols)} columnas históricas de cuota')

# ── 2b. Índice nombre → deudores (para búsqueda alternativa) ─────────────────
nombre_index = {}  # palabra_clave_upper -> list of (sheet_name, row, nombre_completo)
for sheet_name, cfg in SHEETS_CFG.items():
    ws_data = wb_deu_data[sheet_name]
    for r in range(cfg['data_start'], ws_data.max_row + 1):
        nombre = ws_data.cell(r, cfg['nombre_col']).value
        if not nombre:
            continue
        for palabra in normalize_str(str(nombre)).split():
            p_clean = re.sub(r'[^A-Z0-9]', '', palabra)
            if len(p_clean) >= 4:
                nombre_index.setdefault(p_clean, []).append((sheet_name, r, nombre))

# ── 2c. Índice CUIT → nombre desde hojas anteriores de imputaciones ───────────
# Busca el CUIT en hojas previas para obtener el nombre del cliente
wb_imp_ref = openpyxl.load_workbook(IMP_FILE)
all_sheets = wb_imp_ref.sheetnames

# Hojas anteriores a IMP_SHEET (en orden inverso: la más reciente primero)
try:
    idx_actual = all_sheets.index(IMP_SHEET)
    hojas_previas = list(reversed(all_sheets[:idx_actual]))
except ValueError:
    hojas_previas = []

cuit_to_nombre_previo = {}   # cuit -> nombre limpio
cuit_to_cuota_previo  = {}   # cuit -> último número de cuota (int) desde semanas anteriores
for hoja in hojas_previas[:8]:  # revisar hasta 8 semanas anteriores
    try:
        ws_prev = wb_imp_ref[hoja]
    except Exception:
        continue
    for row in ws_prev.iter_rows(min_row=4, max_row=ws_prev.max_row):
        concepto = row[2].value if len(row) > 2 else None
        col_h    = row[7].value if len(row) > 7 else None
        if not concepto or not col_h:
            continue
        col_h_str = str(col_h).strip()
        if not col_h_str or 'PAGO MENOS' in col_h_str or 'Saldo' in col_h_str:
            continue
        cuit = extract_cuit_from_concepto(concepto)
        if cuit and cuit not in cuit_to_nombre_previo:
            # Buscar cuota en cualquier posición (acepta "c 11" con espacio)
            m_cuota = re.search(r'\bc\s*(\d+)', col_h_str, re.IGNORECASE)
            if m_cuota:
                cuit_to_cuota_previo[cuit] = int(m_cuota.group(1))
            # Limpiar nombre: quitar cuota y todo lo que venga después
            nombre_prev = re.sub(r'\s+c\s*\d+.*$', '', col_h_str, flags=re.IGNORECASE).strip()
            # Quitar "parte de ..." si quedó
            nombre_prev = re.sub(r'\s+parte\s+de.*$', '', nombre_prev, flags=re.IGNORECASE).strip()
            # Si hay "y" (dos personas), usar solo la primera
            nombre_prev = re.split(r'\s+y\s+', nombre_prev, maxsplit=1, flags=re.IGNORECASE)[0].strip()
            if nombre_prev:
                cuit_to_nombre_previo[cuit] = nombre_prev

print(f'  -> {len(cuit_to_nombre_previo)} CUITs con nombre desde hojas anteriores')

# ── 2d. Cache de comprobantes del bot (Google Sheets) ────────────────────────
cuit_to_nombre_comprobantes = _cargar_comprobantes()
print(f'  -> {len(cuit_to_nombre_comprobantes)} CUITs en cache de comprobantes')

def buscar_en_deudores_por_nombre(nombre_str):
    """Busca en deudores por palabras del nombre. Devuelve lista de (sheet, row, nombre)."""
    nombre_str = re.split(r'\s+y\s+', nombre_str, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    palabras = [re.sub(r'[^A-Z0-9]', '', p) for p in normalize_str(nombre_str).split()]
    palabras = [p for p in palabras if len(p) >= 4]
    if not palabras:
        return []

    # 1. Intersección estricta: todas las palabras deben coincidir
    sets = [set((s, r) for s, r, _ in nombre_index.get(p, [])) for p in palabras]
    comunes = sets[0]
    for s in sets[1:]:
        comunes &= s

    if not comunes:
        # 2. Mayoría de palabras (n-1 de n)
        from collections import Counter
        cnt = Counter()
        for p in palabras:
            for s, r, _ in nombre_index.get(p, []):
                cnt[(s, r)] += 1
        max_hits = max(cnt.values()) if cnt else 0
        comunes = {k for k, v in cnt.items() if v == max_hits and v >= max(2, len(palabras) - 1)}

    if not comunes:
        # 3. Apellido único: palabra en 1 sola fila + al menos otra palabra confirma esa fila
        for p in sorted(palabras, key=len, reverse=True):
            candidatos = set((s, r) for s, r, _ in nombre_index.get(p, []))
            if len(candidatos) == 1:
                (s_c, r_c) = next(iter(candidatos))
                otras = [q for q in palabras if q != p]
                if any((s_c, r_c) in {(s2, r2) for s2, r2, _ in nombre_index.get(q, [])} for q in otras):
                    comunes = candidatos
                    break

    resultado = []
    seen = set()
    for p in palabras:
        for s, r, n in nombre_index.get(p, []):
            if (s, r) in comunes and (s, r) not in seen:
                resultado.append((s, r, n))
                seen.add((s, r))
    return resultado

# ── 3. Leer S 120 ────────────────────────────────────────────────────────────
print(f'\nCargando imputaciones ({IMP_SHEET})...')
wb_imp = openpyxl.load_workbook(IMP_FILE)
ws120  = wb_imp[IMP_SHEET]

# ── 4. Procesar cada fila pendiente ─────────────────────────────────────────
results   = []   # lista de acciones a tomar
ambiguous = []   # casos reportados para revisión manual
pago_menos= []   # casos de pago insuficiente
written_deu_rows = set()  # (sheet, row) ya usados en esta corrida para evitar doble escritura

for row in ws120.iter_rows(min_row=4, max_row=MAX_ROW):
    fecha_val  = row[0].value  # col A
    monto_val  = row[5].value  # col F (Crédito)
    concepto   = row[2].value  # col C
    col_h_val  = row[7].value  # col H (notas/nombre)

    # Saltar filas vacías
    if fecha_val is None and monto_val is None:
        continue

    row_num = row[0].row

    # Saltar filas ya marcadas en amarillo
    if is_row_yellow(ws120, row_num):
        continue

    # Saltar filas donde H ya tiene "PAGO MENOS" o "Saldo Disponible"
    if col_h_val and ('PAGO MENOS' in str(col_h_val) or 'Saldo Disponible' in str(col_h_val)):
        continue

    # Extraer CUIT
    cuit_raw = extract_cuit_from_concepto(concepto)
    if not cuit_raw:
        ambiguous.append({
            'row': row_num,
            'motivo': 'Sin CUIT extraíble',
            'concepto': concepto,
            'monto': monto_val,
            'fecha': fecha_val,
        })
        continue

    # Buscar en índice por CUIT
    matches = cuit_index.get(cuit_raw, [])
    if not matches:
        # Fallback: buscar nombre en hojas anteriores y luego por nombre en deudores
        nombre_previo = cuit_to_nombre_previo.get(cuit_raw)
        if nombre_previo:
            matches_nombre = buscar_en_deudores_por_nombre(nombre_previo)
            if matches_nombre:
                matches = matches_nombre
                tag = 'fallback nombre' if len(matches) == 1 else f'fallback nombre ({len(matches)} candidatos, desambiguando)'
                print(f'    [{tag}] Fila {row_num}: CUIT {cuit_raw} -> "{nombre_previo}" -> {matches[0][2]}')
            else:
                ambiguous.append({
                    'row': row_num,
                    'motivo': f'CUIT {cuit_raw} no en deudores; nombre "{nombre_previo}" no encontrado en deudores',
                    'concepto': concepto,
                    'monto': monto_val,
                    'fecha': fecha_val,
                })
                continue
        else:
            # Fallback: buscar en cache de comprobantes del bot
            nombre_comprobante = cuit_to_nombre_comprobantes.get(cuit_raw)
            if nombre_comprobante:
                matches_nombre = buscar_en_deudores_por_nombre(nombre_comprobante)
                if matches_nombre:
                    matches = matches_nombre
                    tag = 'fallback comprobantes' if len(matches) == 1 else f'fallback comprobantes ({len(matches)} candidatos, desambiguando)'
                    print(f'    [{tag}] Fila {row_num}: CUIT {cuit_raw} -> "{nombre_comprobante}" -> {matches[0][2]}')
                else:
                    ambiguous.append({
                        'row': row_num,
                        'motivo': f'CUIT {cuit_raw} en comprobantes como "{nombre_comprobante}"; no encontrado en deudores',
                        'concepto': concepto,
                        'monto': monto_val,
                        'fecha': fecha_val,
                    })
                    continue
            else:
                ambiguous.append({
                    'row': row_num,
                    'motivo': f'CUIT {cuit_raw} no encontrado en deudores, semanas anteriores ni comprobantes',
                    'concepto': concepto,
                    'monto': monto_val,
                    'fecha': fecha_val,
                })
                continue

    # Si hay más de un match, intentar desambiguar por monto teórico
    if len(matches) > 1:
        # Obtener teóricos de cada match
        candidatos = []
        for (sname, srow, snombre) in matches:
            cfg = SHEETS_CFG[sname]
            ws_d = wb_deu_data[sname]
            teo  = ws_d.cell(srow, cfg['teo_col']).value
            real = ws_d.cell(srow, cfg['real_col']).value
            candidatos.append((sname, srow, snombre, teo, real))

        # Filtrar ya imputados en mayo 26
        sin_imputar = [(s,r,n,t,rv) for s,r,n,t,rv in candidatos if rv is None]
        if len(sin_imputar) == 0:
            ambiguous.append({
                'row': row_num,
                'motivo': 'Todos los matches ya tienen Mayo 26 imputado',
                'cuit': cuit_raw,
                'monto': monto_val,
                'matches': [(n,t) for _,_,n,t,_ in candidatos],
            })
            continue
        if len(sin_imputar) > 1:
            # Intentar por monto más cercano
            mejor = min(sin_imputar, key=lambda x: abs((x[3] or 0) - (monto_val or 0)))
            diferencia_mejor = abs((mejor[3] or 0) - (monto_val or 0))
            # Verificar que no haya empate
            otros_con_mismo_monto = [
                x for x in sin_imputar
                if x != mejor and abs(abs((x[3] or 0) - (monto_val or 0)) - diferencia_mejor) < 1
            ]
            if otros_con_mismo_monto:
                # Mismo monto en varios lotes: asignar el primero disponible (no usado aún en este run)
                disponibles = [x for x in sin_imputar if (x[0], x[1]) not in written_deu_rows]
                if not disponibles:
                    ambiguous.append({
                        'row': row_num,
                        'motivo': 'Todos los lotes con mismo CUIT ya asignados en este run',
                        'cuit': cuit_raw,
                        'monto': monto_val,
                        'matches': [(n, t) for _,_,n,t,_ in sin_imputar],
                    })
                    continue
                selected = sorted(disponibles, key=lambda x: (x[0], x[1]))[0]
            else:
                selected = mejor
        else:
            selected = sin_imputar[0]
    else:
        sname, srow, snombre = matches[0]
        cfg   = SHEETS_CFG[sname]
        ws_d  = wb_deu_data[sname]
        teo   = ws_d.cell(srow, cfg['teo_col']).value
        real  = ws_d.cell(srow, cfg['real_col']).value
        selected = (sname, srow, snombre, teo, real)

    sname, srow, snombre, teo_val, real_existente = selected
    cfg = SHEETS_CFG[sname]

    # Si ya está imputado en mayo 26 O si esta corrida ya lo procesó
    if real_existente is not None or (sname, srow) in written_deu_rows:
        ambiguous.append({
            'row': row_num,
            'motivo': f'Mayo 26 ya imputado ({real_existente}) o destino duplicado en {sname} fila {srow}',
            'cliente': snombre,
            'cuit': cuit_raw,
            'monto': monto_val,
        })
        continue

    # Comparar monto transferido vs teórico
    monto_num = monto_val if isinstance(monto_val, (int, float)) else 0
    teo_num   = teo_val   if isinstance(teo_val,  (int, float)) else 0

    if monto_num < teo_num and (teo_num - monto_num) > TOLERANCE:
        pago_menos.append({
            'row': row_num,
            'cliente': snombre,
            'cuit': cuit_raw,
            'transferido': monto_num,
            'teorico': round(teo_num),
            'diferencia': round(teo_num - monto_num),
        })
        continue

    # Determinar número de cuota
    ws_d = wb_deu_data[sname]
    cuota_col_val = ws_d.cell(srow, cfg['cuota_col']).value

    # Si la cuota actual dice "parte de X", el siguiente pago completo es X+1
    if isinstance(cuota_col_val, str) and 'parte' in cuota_col_val.lower():
        m = re.search(r'\d+', cuota_col_val)
        if m:
            next_cuota = int(m.group()) + 1
        else:
            ambiguous.append({
                'row': row_num,
                'motivo': f'Cuota dice "parte de..." pero no se pudo extraer número ({cuota_col_val!r})',
                'cliente': snombre,
                'cuit': cuit_raw,
                'hoja': sname,
                'hoja_fila': srow,
            })
            continue
    else:
        # Buscar el número de cuota más alto en las columnas históricas de la fila
        hist_cols = cuota_history_cols.get(sname, [])
        max_hist = None
        for hc in hist_cols:
            n_celda = max_cuota_celda(ws_d.cell(srow, hc).value)
            if n_celda is not None and (max_hist is None or n_celda > max_hist):
                max_hist = n_celda

        if max_hist is not None:
            next_cuota = max_hist + 1
        elif cuit_raw in cuit_to_cuota_previo:
            next_cuota = cuit_to_cuota_previo[cuit_raw] + 1
            print(f'    [cuota imputaciones] Fila {row_num}: usando cuota previa {cuit_to_cuota_previo[cuit_raw]}+1={next_cuota}')
        elif cuit_raw in CUOTA_OVERRIDE:
            next_cuota = CUOTA_OVERRIDE[cuit_raw]
            print(f'    [cuota override] Fila {row_num}: usando cuota manual {next_cuota} para {snombre}')
        else:
            ambiguous.append({
                'row': row_num,
                'motivo': 'No se pudo determinar número de cuota (sin historial en deudores ni en imputaciones)',
                'cliente': snombre,
                'cuit': cuit_raw,
                'hoja': sname,
                'hoja_fila': srow,
            })
            continue

    fecha_dt = parse_date(fecha_val)

    # Determinar si hay que incluir el número de lote en el label
    # (cuando hay múltiples filas en deudores con el mismo nombre para este CUIT)
    todos_matches = cuit_index.get(cuit_raw, []) + [
        (s, r, n) for s, r, n in buscar_en_deudores_por_nombre(cuit_to_nombre_previo.get(cuit_raw, ''))
        if cuit_raw not in cuit_index
    ] if cuit_raw not in cuit_index else cuit_index.get(cuit_raw, [])
    nombres_distintos = {str(n).strip().upper() for _, _, n in todos_matches}
    if len(todos_matches) > 1 and len(nombres_distintos) == 1:
        # Mismo nombre en todos los lotes → agregar número de lote
        lote_val = wb_deu_data[sname].cell(srow, cfg['lote_col']).value
        lote_str = f' l{lote_val}' if lote_val is not None else ''
    else:
        lote_str = ''

    written_deu_rows.add((sname, srow))
    results.append({
        'imp_row': row_num,
        'cuit': cuit_raw,
        'cliente': snombre,
        'lote_str': lote_str,
        'hoja': sname,
        'hoja_fila': srow,
        'monto_real': monto_num,
        'monto_teo': round(teo_num),
        'diferencia': round(monto_num - teo_num),
        'cuota': next_cuota,
        'fecha': fecha_dt,
    })

# ── 5. Reporte ───────────────────────────────────────────────────────────────
print('\n' + '='*70)
print(f'  RESULTADOS DRY-RUN  (DRY_RUN={DRY_RUN})')
print('='*70)

print(f'\nOK PARA IMPUTAR ({len(results)}):')
for r in results:
    dif_str = f'+{r["diferencia"]}' if r['diferencia'] >= 0 else str(r['diferencia'])
    lote_disp = r.get('lote_str', '')
    print(f'  Fila {r["imp_row"]:3d}: {r["cliente"][:35]:<35} | '
          f'CUIT {r["cuit"]} | '
          f'${r["monto_real"]:>10,.0f} (teo ${r["monto_teo"]:>10,.0f}, dif {dif_str:>8}) | '
          f'{lote_disp}c{r["cuota"]} | {r["fecha"].strftime("%d/%m/%Y") if r["fecha"] else "?"} | '
          f'{r["hoja"]} fila {r["hoja_fila"]}')

print(f'\nPAGO MENOS ({len(pago_menos)}) -- se escribira "PAGO MENOS" en col H:')
for p in pago_menos:
    print(f'  Fila {p["row"]:3d}: {p["cliente"][:35]:<35} | '
          f'CUIT {p["cuit"]} | '
          f'${p["transferido"]:>10,.0f} vs teo ${p["teorico"]:>10,.0f} | '
          f'diferencia ${p["diferencia"]:>8,.0f}')

print(f'\nCASOS AMBIGUOS ({len(ambiguous)}) -- sin imputar, revisar manualmente:')
for a in ambiguous:
    motivo = a.get('motivo', '')
    fila   = a.get('row', '?')
    cliente= a.get('cliente', a.get('concepto', ''))
    monto  = a.get('monto', '')
    matches= a.get('matches', '')
    print(f'  Fila {fila:3d}: {motivo}')
    if cliente:
        print(f'         Cliente/Concepto: {str(cliente)[:60]}')
    if monto:
        print(f'         Monto: ${monto:>10,.0f}' if isinstance(monto, (int,float)) else f'         Monto: {monto}')
    if matches:
        print(f'         Candidatos: {matches}')

print('\n' + '='*70)
print(f'  Total filas {IMP_SHEET} a procesar: {len(results)+len(pago_menos)+len(ambiguous)}')
print(f'  -> Imputar: {len(results)} | PAGO MENOS: {len(pago_menos)} | Revisar: {len(ambiguous)}')
print('='*70)

# ── 6. Escritura (solo si DRY_RUN=False) ────────────────────────────────────
if DRY_RUN:
    print('\n[DRY RUN] No se modifico ningun archivo.')
    print('  Cambia DRY_RUN = False y volvé a ejecutar para aplicar.')
else:
    # Backup
    for path in [IMP_FILE, DEU_FILE]:
        bak = path + '.bak'
        if not os.path.exists(bak):
            shutil.copy2(path, bak)
            print(f'Backup: {bak}')

    ws120_edit = wb_imp[IMP_SHEET]

    # Escribir PAGO MENOS en imputaciones
    for p in pago_menos:
        row_num = p['row']
        ws120_edit.cell(row_num, 8).value = 'PAGO MENOS'   # col H

    # Escribir imputaciones
    for r in results:
        row_num  = r['imp_row']
        sname    = r['hoja']
        srow     = r['hoja_fila']
        cfg      = SHEETS_CFG[sname]
        ws_edit  = wb_deu_edit[sname]

        # --- Deudores ---
        ws_edit.cell(srow, cfg['real_col']).value  = r['monto_real']
        ws_edit.cell(srow, cfg['cuota_col']).value = r['cuota']
        ws_edit.cell(srow, cfg['fecha_col']).value = r['fecha']

        # --- Imputaciones: col H (nombre [l{lote}] c{cuota}), amarillo ---
        nombre_corto = str(r['cliente'])[:38]
        ws120_edit.cell(row_num, 8).value = f"{nombre_corto}{r['lote_str']} c{r['cuota']}"
        for cell in ws120_edit[row_num]:
            cell.fill = YELLOW_FILL

    wb_deu_edit.save(DEU_FILE)
    wb_imp.save(IMP_FILE)
    print(f'\nArchivos guardados. Imputadas: {len(results)} | PAGO MENOS: {len(pago_menos)}')
